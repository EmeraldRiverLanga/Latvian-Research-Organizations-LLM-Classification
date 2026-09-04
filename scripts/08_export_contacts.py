"""Export the organization list and contacts to a two-sheet workbook."""
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ORGS = "data/processed/07_with_activity.csv"
CONTACTS = "data/processed/07_contacts.jsonl"
DST = "output/zinatnisko_biedribu_datubaze.xlsx"

HEADER = PatternFill("solid", fgColor="D9D9D9")

ORG_COLUMNS = {
    "Reģ. nr.": 13, "Nosaukums": 48, "Klase": 30, "Grupa": 30,
    "aktivitate": 18, "pedejais_parskats": 12, "darbinieki": 11,
    "Juridiskā forma": 14, "Adrese": 34, "Deklarētās jomas": 28,
}

CONTACT_COLUMNS = {
    "Reģ. nr.": 13, "Nosaukums": 44, "statuss": 20, "e_pasts": 28,
    "telefons": 14, "majaslapa": 40, "facebook": 34, "youtube": 26,
    "linkedin": 26, "apraksts": 55, "atrasts_katalogos": 16, "parbaudits": 12,
}

GROUP = {
    "Pētnieku biedrība": "Pētnieku biedrības",
    "Medicīnas specialitāte": "Profesionālas zinātnes asociācijas",
    "Aprūpes un veselības profesija": "Profesionālas zinātnes asociācijas",
    "Praktizējoša psiholoģija vai terapija": "Profesionālas zinātnes asociācijas",
    "Inženierzinātņu profesija": "Profesionālas zinātnes asociācijas",
    "Juridiskā profesija": "Profesionālas zinātnes asociācijas",
    "Cita akadēmiska profesija": "Profesionālas zinātnes asociācijas",
    "Pētniecības institūts vai centrs": "Ar zinātni saistītās organizācijas",
    "Studentu un jauno zinātnieku biedrība": "Ar zinātni saistītās organizācijas",
    "Institūciju asociācija": "Ar zinātni saistītās organizācijas",
    "Skolotāju priekšmetu asociācija": "Ar zinātni saistītās organizācijas",
    "Vēsture un novadpētniecība": "Ar zinātni saistītās organizācijas",
    "Popularizēšana un izglītība": "Ar zinātni saistītās organizācijas",
    "Neskaidrs": "Ar zinātni saistītās organizācijas",
}

os.makedirs("output", exist_ok=True)

orgs = pd.read_csv(ORGS, encoding="utf-8", dtype={"Reģ. nr.": str})
orgs["Grupa"] = orgs["Klase"].map(GROUP)
orgs = orgs[[c for c in ORG_COLUMNS if c in orgs.columns]]
orgs = orgs.sort_values(["Klase", "Nosaukums"])

if os.path.exists(CONTACTS):
    con = pd.read_json(CONTACTS, lines=True, dtype={"regcode": str})
    con = con.rename(columns={"regcode": "Reģ. nr.", "name": "Nosaukums"})
    con = con[[c for c in CONTACT_COLUMNS if c in con.columns]]
    con = con.sort_values(["statuss", "Nosaukums"])
else:
    con = pd.DataFrame(columns=list(CONTACT_COLUMNS))

with pd.ExcelWriter(DST, engine="openpyxl") as w:
    orgs.to_excel(w, sheet_name="Organizācijas", index=False)
    con.to_excel(w, sheet_name="Kontakti", index=False)

    for sheet, widths in (("Organizācijas", ORG_COLUMNS),
                          ("Kontakti", CONTACT_COLUMNS)):
        ws = w.sheets[sheet]
        for row in ws.iter_rows():
            for c in row:
                c.font = Font(name="Arial", size=10)
                c.alignment = Alignment(vertical="top", wrap_text=True)
        for c in ws[1]:
            c.font = Font(name="Arial", size=10, bold=True)
            c.fill = HEADER
        for i, name in enumerate([c.value for c in ws[1]], start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(name, 15)
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

print(f"Written to {DST}")
print(f"  Organizācijas: {len(orgs)} rows")
print(f"  Kontakti: {len(con)} rows")