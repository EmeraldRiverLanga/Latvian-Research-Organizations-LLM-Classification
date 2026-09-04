"""Export the review queue and candidates to a workbook for manual review."""
import os

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

CANDIDATES = "data/processed/04_candidates.csv"
REVIEW = "data/processed/04_review_queue.csv"
DST = "output/review.xlsx"

COLUMNS = ["regcode", "name", "level", "confidence", "reason",
           "type_text", "declared_areas", "address"]

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")

WIDTHS = {"regcode": 13, "name": 45, "level": 7, "confidence": 11,
          "reason": 55, "type_text": 14, "declared_areas": 35,
          "address": 35, "review_group": 30, "manual_level": 13, "note": 30}

os.makedirs("output", exist_ok=True)

candidates = pd.read_csv(CANDIDATES, encoding="utf-8", dtype={"regcode": str})
review = pd.read_csv(REVIEW, encoding="utf-8", dtype={"regcode": str})


def prepare(df, extra=None):
    """Order the columns and append the two manual-input columns."""
    cols = (["review_group"] if extra else []) + COLUMNS
    out = df[[c for c in cols if c in df.columns]].copy()
    out["manual_level"] = ""
    out["note"] = ""
    return out


sheets = {
    "Review queue": prepare(review, extra=True),
    "Level 1": prepare(candidates[candidates["level"] == 1]),
    "Level 2": prepare(candidates[candidates["level"] == 2]),
    "Level 3": prepare(candidates[candidates["level"] == 3]),
}

with pd.ExcelWriter(DST, engine="openpyxl") as writer:
    for title, frame in sheets.items():
        frame.to_excel(writer, sheet_name=title, index=False)
        ws = writer.sheets[title]

        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Arial", size=10)
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        for cell in ws[1]:
            cell.font = Font(name="Arial", size=10, bold=True)
            cell.fill = HEADER_FILL

        for i, name in enumerate(frame.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 15)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Dropdown on manual_level; blank means "the classification stands".
        col = get_column_letter(list(frame.columns).index("manual_level") + 1)
        note = get_column_letter(list(frame.columns).index("note") + 1)
        last = len(frame) + 1

        dv = DataValidation(type="list", formula1='"0,1,2,3"', allow_blank=True)
        dv.error = "Enter 0, 1, 2 or 3, or leave the cell empty."
        ws.add_data_validation(dv)
        dv.add(f"{col}2:{col}{last}")

        for r in range(2, last + 1):
            ws[f"{col}{r}"].fill = INPUT_FILL
            ws[f"{note}{r}"].fill = INPUT_FILL

print(f"Written to {DST}")
for title, frame in sheets.items():
    print(f"  {title}: {len(frame)} rows")
print("\nFill manual_level only where you disagree; an empty cell keeps the "
      "classifier's level. The note column is free text.")